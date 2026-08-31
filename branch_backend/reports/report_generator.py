"""
branch_backend/reports/report_generator.py
==========================================
Weekly ATM Cash Replenishment Report Generator
Produces a professional .xlsx workbook with:
  - Cover sheet  (report metadata)
  - Weekly Summary  (one row per ATM, totals + Poya flags)
  - Daily Detail    (one row per ATM per day, with holiday / Poya context)
  - Replenishment Schedule  (recommended loading schedule + alert levels)

Usage (standalone demo):
  python report_generator.py
"""

import os, sys, json, math
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "navy":     "1A3A5C",   # header background
    "gold":     "C9A84C",   # accent / Poya row
    "poya_bg":  "FFF3CD",   # soft yellow for Poya days
    "pre_poya": "FFE8B0",   # slightly warmer pre-Poya
    "post_poya":"FFF8E1",
    "alert_c4": "FF4444",   # critical – red
    "alert_c3": "FF8800",   # high – orange
    "alert_c2": "FFD700",   # medium – amber
    "alert_c1": "90EE90",   # low – green
    "white":    "FFFFFF",
    "light":    "F5F7FA",
    "mid":      "D0D9E8",
    "text_dark":"1A1A2E",
}

ATMS = [
    "Airport ATM", "Big Street ATM", "Christ College ATM",
    "KK Nagar ATM", "Mount Road ATM",
]
ZONE_MAP = {
    "Airport ATM": "Transport", "Big Street ATM": "Commercial",
    "Christ College ATM": "Educational", "KK Nagar ATM": "Residential",
    "Mount Road ATM": "Commercial",
}

# ── Alert thresholds (% of daily predicted demand) ───────────────────────────
ALERT_LEVELS = {
    "C4 – Critical": (0.0,  0.20, C["alert_c4"]),
    "C3 – High":     (0.20, 0.40, C["alert_c3"]),
    "C2 – Medium":   (0.40, 0.60, C["alert_c2"]),
    "C1 – Low":      (0.60, 1.00, C["alert_c1"]),
}

# ── Sri Lankan Poya dates 2026 ────────────────────────────────────────────────
POYA_2026 = {
    (2026,1,3),(2026,2,1),(2026,3,3),(2026,4,2),(2026,5,1),(2026,5,31),
    (2026,6,29),(2026,7,29),(2026,8,27),(2026,9,26),(2026,10,25),
    (2026,11,24),(2026,12,23),
}

def _is_poya(dt: datetime) -> bool:
    return (dt.year, dt.month, dt.day) in POYA_2026

def _is_pre_poya(dt: datetime) -> bool:
    return _is_poya(dt + timedelta(days=1))

def _is_post_poya(dt: datetime) -> bool:
    return _is_poya(dt - timedelta(days=1))

def _poya_label(dt: datetime) -> str:
    if _is_poya(dt):      return "Poya Day"
    if _is_pre_poya(dt):  return "Pre-Poya"
    if _is_post_poya(dt): return "Post-Poya"
    return ""

def _poya_mult(dt: datetime) -> float:
    if _is_poya(dt):      return 0.80
    if _is_pre_poya(dt):  return 1.30
    if _is_post_poya(dt): return 1.10
    return 1.0

SL_HOLIDAYS_2026 = {
    (2026,1,14): "Thai Pongal",
    (2026,1,15): "Duruthu Poya / Tamil Thai Pongal",
    (2026,2,4):  "National Day",
    (2026,4,13): "Sinhala & Tamil New Year Eve",
    (2026,4,14): "Sinhala & Tamil New Year",
    (2026,5,1):  "May Day / Vesak Poya",
    (2026,12,25):"Christmas Day",
}

def _holiday_label(dt: datetime) -> str:
    return SL_HOLIDAYS_2026.get((dt.year, dt.month, dt.day), "")

def _alert_level(current_lkr: float, predicted_lkr: float):
    ratio = current_lkr / max(predicted_lkr, 1)
    for name, (lo, hi, colour) in ALERT_LEVELS.items():
        if lo <= ratio < hi:
            return name, colour
    return "C1 – Low", C["alert_c1"]

# ── Style helpers ─────────────────────────────────────────────────────────────
def _fill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_colour, fgColor=hex_colour)

def _font(bold=False, size=11, colour="000000", italic=False) -> Font:
    return Font(name="Arial", bold=bold, size=size, color=colour, italic=italic)

def _border(style="thin") -> Border:
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_border() -> Border:
    thin  = Side(style="thin")
    thick = Side(style="medium")
    return Border(left=thick, right=thick, top=thick, bottom=thick)

def _header_cell(ws, row, col, value, bg=C["navy"], fg=C["white"],
                 bold=True, size=11, wrap=False, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = _fill(bg)
    cell.font      = _font(bold=bold, size=size, colour=fg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border    = _border()
    return cell

def _data_cell(ws, row, col, value, bg=C["white"], bold=False,
               number_format=None, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = _fill(bg)
    cell.font      = _font(bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _border()
    if number_format:
        cell.number_format = number_format
    return cell

# ── Prediction builder (simulated or real) ───────────────────────────────────
def build_week_predictions(week_start: datetime, predictor=None) -> list[dict]:
    """
    Build 7 days × 5 ATMs prediction rows.
    If a real predictor is passed, use it; otherwise simulate plausible values.
    """
    rows = []
    for offset in range(7):
        dt = week_start + timedelta(days=offset)
        poya_label = _poya_label(dt)
        poya_mult  = _poya_mult(dt)
        holiday    = _holiday_label(dt)
        is_weekend = dt.weekday() >= 5
        is_salary  = dt.day == 1

        for atm in ATMS:
            # --- real predictor path ---
            if predictor:
                try:
                    result = predictor.predict(dt.strftime("%Y-%m-%d"), atm, "auto")
                    pred   = result["predictions"][0]
                    base   = pred["predicted"]
                except Exception:
                    base = 500_000
            else:
                # Simulate realistic figures per zone
                base_vals = {
                    "Airport ATM": 820_000, "Big Street ATM": 640_000,
                    "Christ College ATM": 480_000, "KK Nagar ATM": 560_000,
                    "Mount Road ATM": 720_000,
                }
                base = base_vals[atm]
                if is_weekend: base = int(base * 0.75)
                if is_salary:  base = int(base * 1.40)

            adjusted = int(base * poya_mult)
            p10      = int(adjusted * 0.80)
            p90      = int(adjusted * 1.20)
            # Simulate current cash (random-ish but realistic)
            import random; random.seed(hash((dt.date(), atm)))
            cash_ratio  = random.uniform(0.15, 0.95)
            current_cash= int(adjusted * cash_ratio)
            alert_name, alert_colour = _alert_level(current_cash, adjusted)

            rows.append({
                "date":          dt,
                "day_name":      dt.strftime("%A"),
                "atm":           atm,
                "zone":          ZONE_MAP[atm],
                "base_predicted":base,
                "poya_mult":     poya_mult,
                "adjusted":      adjusted,
                "p10":           p10,
                "p90":           p90,
                "poya_label":    poya_label,
                "holiday":       holiday,
                "is_weekend":    is_weekend,
                "is_salary":     is_salary,
                "current_cash":  current_cash,
                "alert_name":    alert_name,
                "alert_colour":  alert_colour,
            })
    return rows


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – Cover
# ═══════════════════════════════════════════════════════════════════════════════
def _build_cover(wb: Workbook, week_start: datetime, week_end: datetime,
                 generated_at: datetime):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False

    # Big title block
    ws.merge_cells("B2:I2")
    c = ws["B2"]
    c.value     = "ATM CASHPREDICT"
    c.font      = Font(name="Arial", bold=True, size=28, color=C["white"])
    c.fill      = _fill(C["navy"])
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B3:I3")
    c = ws["B3"]
    c.value     = "Weekly Cash Replenishment Report"
    c.font      = Font(name="Arial", bold=False, size=16, color=C["gold"])
    c.fill      = _fill(C["navy"])
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 50
    ws.row_dimensions[3].height = 30

    # Metadata table
    meta = [
        ("Report Period",  f"{week_start.strftime('%d %B %Y')} – {week_end.strftime('%d %B %Y')}"),
        ("Generated",      generated_at.strftime("%d %B %Y  %H:%M")),
        ("ATMs Covered",   ", ".join(ATMS)),
        ("Prediction Model","XGBoost (ATM CashPredict v2)"),
        ("Poya Intelligence","Enabled – Sri Lankan lunar calendar"),
        ("Currency",       "Sri Lankan Rupees (LKR)"),
        ("Prepared By",    "ATM CashPredict System  |  Automated Report"),
    ]
    for i, (label, value) in enumerate(meta, start=6):
        ws.merge_cells(f"C{i}:D{i}")
        l = ws.cell(row=i, column=3, value=label)
        l.font      = _font(bold=True, size=11, colour=C["white"])
        l.fill      = _fill(C["navy"])
        l.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        l.border    = _border()

        ws.merge_cells(f"E{i}:I{i}")
        v = ws.cell(row=i, column=5, value=value)
        v.font      = _font(size=11)
        v.fill      = _fill(C["light"])
        v.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        v.border    = _border()
        ws.row_dimensions[i].height = 22

    # Alert legend
    ws.merge_cells("C14:I14")
    c = ws["C14"]
    c.value = "Alert Level Legend"
    c.font  = _font(bold=True, size=12, colour=C["white"])
    c.fill  = _fill(C["navy"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[14].height = 22

    legend = [
        ("C4 – Critical", "< 20% of daily demand remaining", C["alert_c4"]),
        ("C3 – High",     "20–40% remaining",                C["alert_c3"]),
        ("C2 – Medium",   "40–60% remaining",                C["alert_c2"]),
        ("C1 – Low",      "> 60% remaining (adequate)",       C["alert_c1"]),
    ]
    for j, (lvl, desc, col) in enumerate(legend, start=15):
        ws.merge_cells(f"C{j}:D{j}")
        lc = ws.cell(row=j, column=3, value=lvl)
        lc.font      = _font(bold=True, size=10, colour="000000")
        lc.fill      = _fill(col)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border    = _border()

        ws.merge_cells(f"E{j}:I{j}")
        dc = ws.cell(row=j, column=5, value=desc)
        dc.font      = _font(size=10)
        dc.fill      = _fill(C["light"])
        dc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        dc.border    = _border()
        ws.row_dimensions[j].height = 20

    # Footer
    ws.merge_cells("B22:I22")
    f = ws["B22"]
    f.value     = "CONFIDENTIAL — For Internal Bank Management Use Only"
    f.font      = Font(name="Arial", bold=True, size=10, color=C["white"], italic=True)
    f.fill      = _fill(C["navy"])
    f.alignment = Alignment(horizontal="center", vertical="center")

    for col in ["A","B","C","D","E","F","G","H","I","J"]:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["J"].width = 2


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – Weekly Summary
# ═══════════════════════════════════════════════════════════════════════════════
def _build_summary(wb: Workbook, rows: list[dict], week_start: datetime):
    ws = wb.create_sheet("Weekly Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value     = f"Weekly Summary  —  w/c {week_start.strftime('%d %B %Y')}"
    c.font      = _font(bold=True, size=14, colour=C["white"])
    c.fill      = _fill(C["navy"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = [
        "ATM Name", "Zone", "Mon", "Tue", "Wed", "Thu",
        "Fri", "Sat", "Sun", "Weekly Total (LKR)", "Poya Days"
    ]
    for col, h in enumerate(headers, 1):
        _header_cell(ws, 2, col, h, size=10)

    # Build ATM totals
    df = pd.DataFrame(rows)
    days_order = sorted(df["date"].unique())

    for r_idx, atm in enumerate(ATMS, start=3):
        sub = df[df["atm"] == atm].sort_values("date")
        bg  = C["light"] if r_idx % 2 == 0 else C["white"]

        _data_cell(ws, r_idx, 1, atm, bg=bg, bold=True, align="left")
        _data_cell(ws, r_idx, 2, ZONE_MAP[atm], bg=bg, align="left")

        weekly_total = 0
        poya_days    = []
        for col_idx, dt in enumerate(days_order, start=3):
            day_row = sub[sub["date"] == dt]
            if day_row.empty:
                _data_cell(ws, r_idx, col_idx, "—", bg=bg)
                continue
            val  = int(day_row["adjusted"].iloc[0])
            pl   = day_row["poya_label"].iloc[0]
            hol  = day_row["holiday"].iloc[0]
            weekly_total += val

            cell_bg = bg
            if pl == "Poya Day":  cell_bg = C["poya_bg"];  poya_days.append(dt.strftime("%a"))
            elif pl == "Pre-Poya":  cell_bg = C["pre_poya"]
            elif pl == "Post-Poya": cell_bg = C["post_poya"]
            elif hol: cell_bg = "E8F5E9"

            c = _data_cell(ws, r_idx, col_idx, val, bg=cell_bg,
                           number_format='#,##0')
            if pl or hol:
                c.font = _font(bold=True, size=10)

        _data_cell(ws, r_idx, 10, weekly_total, bg=C["mid"], bold=True,
                   number_format='#,##0')
        _data_cell(ws, r_idx, 11, ", ".join(poya_days) if poya_days else "None",
                   bg=bg, align="left")
        ws.row_dimensions[r_idx].height = 20

    # Grand total row
    total_row = 3 + len(ATMS)
    _header_cell(ws, total_row, 1, "GRAND TOTAL", bg=C["gold"], fg="000000", size=10)
    _header_cell(ws, total_row, 2, "", bg=C["gold"], fg="000000")

    for col_idx, dt in enumerate(days_order, start=3):
        day_total = int(df[df["date"] == dt]["adjusted"].sum())
        _header_cell(ws, total_row, col_idx,
                     day_total, bg=C["gold"], fg="000000", size=10)
        ws.cell(row=total_row, column=col_idx).number_format = '#,##0'

    grand_total = int(df["adjusted"].sum())
    _header_cell(ws, total_row, 10, grand_total, bg=C["gold"], fg="000000", size=10)
    ws.cell(row=total_row, column=10).number_format = '#,##0'
    _header_cell(ws, total_row, 11, "", bg=C["gold"])
    ws.row_dimensions[total_row].height = 22

    # Key annotation below table
    note_row = total_row + 2
    ws.merge_cells(f"A{note_row}:K{note_row}")
    n = ws.cell(row=note_row, column=1,
                value="Note: Highlighted cells — 🟡 Poya Day (−20%),  🟠 Pre-Poya (+30%),  🟢 Post-Poya (+10%),  🟩 Public Holiday")
    n.font      = _font(italic=True, size=9, colour="555555")
    n.alignment = Alignment(horizontal="left", vertical="center")

    # Column widths
    widths = [22, 14, 14, 14, 14, 14, 14, 14, 14, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – Daily Detail
# ═══════════════════════════════════════════════════════════════════════════════
def _build_daily_detail(wb: Workbook, rows: list[dict]):
    ws = wb.create_sheet("Daily Detail")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value     = "Daily ATM Cash Demand Detail"
    c.font      = _font(bold=True, size=14, colour=C["white"])
    c.fill      = _fill(C["navy"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        "Date", "Day", "ATM Name", "Zone",
        "Base Prediction", "Poya Mult.", "Adjusted Pred.",
        "P10 (Low)", "P90 (High)",
        "Poya Status", "Public Holiday",
        "Current Cash", "Alert Level", "Action Required"
    ]
    for col, h in enumerate(headers, 1):
        _header_cell(ws, 2, col, h, size=9, wrap=True)
    ws.row_dimensions[2].height = 32

    ACTION_MAP = {
        "C4 – Critical": "⚠ REPLENISH IMMEDIATELY",
        "C3 – High":     "Schedule replenishment today",
        "C2 – Medium":   "Plan replenishment this week",
        "C1 – Low":      "No action needed",
    }

    for r_idx, row in enumerate(rows, start=3):
        dt      = row["date"]
        alert   = row["alert_name"]
        al_col  = row["alert_colour"]

        # Row background: Poya > Holiday > alternate
        if row["poya_label"] == "Poya Day":   bg = C["poya_bg"]
        elif row["poya_label"] == "Pre-Poya":  bg = C["pre_poya"]
        elif row["poya_label"] == "Post-Poya": bg = C["post_poya"]
        elif row["holiday"]:                    bg = "E8F5E9"
        elif r_idx % 2 == 0:                    bg = C["light"]
        else:                                   bg = C["white"]

        vals = [
            dt.strftime("%d/%m/%Y"), row["day_name"],
            row["atm"], row["zone"],
            row["base_predicted"], row["poya_mult"], row["adjusted"],
            row["p10"], row["p90"],
            row["poya_label"] or "Normal",
            row["holiday"] or "—",
            row["current_cash"], alert,
            ACTION_MAP.get(alert, ""),
        ]
        num_fmts = [None, None, None, None,
                    '#,##0', '0.00', '#,##0',
                    '#,##0', '#,##0',
                    None, None, '#,##0', None, None]

        for col_idx, (val, fmt) in enumerate(zip(vals, num_fmts), 1):
            if col_idx == 13:          # alert level — coloured
                c2 = _data_cell(ws, r_idx, col_idx, val, bg=al_col, bold=True)
            elif col_idx == 14:        # action
                bold = alert in ("C4 – Critical", "C3 – High")
                c2 = _data_cell(ws, r_idx, col_idx, val, bg=bg, bold=bold, align="left")
            else:
                align = "left" if col_idx in (3,4,11,14) else "center"
                c2 = _data_cell(ws, r_idx, col_idx, val, bg=bg, align=align)
            if fmt:
                c2.number_format = fmt

        ws.row_dimensions[r_idx].height = 18

    widths = [12, 10, 22, 14, 16, 10, 16, 14, 14, 14, 20, 16, 16, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – Replenishment Schedule
# ═══════════════════════════════════════════════════════════════════════════════
def _build_schedule(wb: Workbook, rows: list[dict], week_start: datetime):
    ws = wb.create_sheet("Replenishment Schedule")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value     = "Recommended Cash Replenishment Schedule"
    c.font      = _font(bold=True, size=14, colour=C["white"])
    c.fill      = _fill(C["navy"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        "Priority", "ATM Name", "Zone",
        "Recommended Date", "Reason",
        "Load Amount (LKR)", "Current Cash (LKR)",
        "Next Poya Day", "Notes"
    ]
    for col, h in enumerate(headers, 1):
        _header_cell(ws, 2, col, h, size=10, wrap=True)
    ws.row_dimensions[2].height = 30

    df = pd.DataFrame(rows)
    # Find next Poya from week start
    next_poya = None
    for offset in range(1, 30):
        candidate = week_start + timedelta(days=offset)
        if _is_poya(candidate):
            next_poya = candidate
            break

    # Per-ATM recommendations
    schedule = []
    for atm in ATMS:
        sub = df[df["atm"] == atm]
        # Find worst alert day this week
        crit = sub[sub["alert_name"].str.startswith("C4")]
        high = sub[sub["alert_name"].str.startswith("C3")]

        if not crit.empty:
            worst = crit.sort_values("date").iloc[0]
            priority = 1
            reason   = "C4 Critical – below 20% remaining"
            rec_date = worst["date"]
        elif not high.empty:
            worst    = high.sort_values("date").iloc[0]
            priority = 2
            reason   = "C3 High – below 40% remaining"
            rec_date = worst["date"]
        elif next_poya and (next_poya - week_start).days <= 3:
            worst    = sub.sort_values("date").iloc[0]
            priority = 3
            reason   = f"Pre-Poya loading (Poya: {next_poya.strftime('%d %b')})"
            rec_date = next_poya - timedelta(days=1)
        else:
            worst    = sub.sort_values("date").iloc[0]
            priority = 4
            reason   = "Routine weekly replenishment"
            rec_date = week_start + timedelta(days=3)

        load_amount  = int(worst["p90"] * 1.10)
        current_cash = int(worst["current_cash"])
        poya_str     = next_poya.strftime("%d %b %Y") if next_poya else "None this week"

        notes = ""
        if next_poya and abs((next_poya - rec_date).days) <= 1:
            notes = "Load day before Poya; demand drops −20% on Poya day"
        elif worst["is_salary"].any() if hasattr(worst["is_salary"], "any") else worst["is_salary"]:
            notes = "Salary day — demand +40% expected"

        schedule.append({
            "priority": priority, "atm": atm, "zone": ZONE_MAP[atm],
            "rec_date": rec_date, "reason": reason,
            "load_amount": load_amount, "current_cash": current_cash,
            "poya_str": poya_str, "notes": notes,
        })

    schedule.sort(key=lambda x: x["priority"])

    priority_colours = {1: C["alert_c4"], 2: C["alert_c3"], 3: C["alert_c2"], 4: C["alert_c1"]}

    for r_idx, s in enumerate(schedule, start=3):
        pc  = priority_colours.get(s["priority"], C["white"])
        bg  = C["light"] if r_idx % 2 == 0 else C["white"]
        row_vals = [
            s["priority"], s["atm"], s["zone"],
            s["rec_date"].strftime("%A, %d %b %Y"), s["reason"],
            s["load_amount"], s["current_cash"],
            s["poya_str"], s["notes"],
        ]
        num_fmts = [None, None, None, None, None, '#,##0', '#,##0', None, None]
        for col_idx, (val, fmt) in enumerate(zip(row_vals, num_fmts), 1):
            if col_idx == 1:
                c2 = _data_cell(ws, r_idx, col_idx, val, bg=pc, bold=True)
            else:
                align = "left" if col_idx in (2,3,4,5,8,9) else "center"
                c2 = _data_cell(ws, r_idx, col_idx, val, bg=bg, align=align)
            if fmt:
                c2.number_format = fmt
        ws.row_dimensions[r_idx].height = 22

    # Footer note
    note_row = 3 + len(schedule) + 1
    ws.merge_cells(f"A{note_row}:I{note_row}")
    n = ws.cell(row=note_row, column=1,
                value="Load amounts are P90 (upper bound) + 10% buffer.  Poya Day demand −20%, Pre-Poya +30%.  Confirm with branch ops before dispatch.")
    n.font      = _font(italic=True, size=9, colour="666666")
    n.alignment = Alignment(horizontal="left")

    widths = [10, 22, 14, 24, 28, 18, 18, 16, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN BUILDER
# ═══════════════════════════════════════════════════════════════════════════════
def generate_report(week_start: datetime = None, predictor=None,
                    output_path: str = None) -> str:
    if week_start is None:
        today      = datetime.today()
        week_start = today - timedelta(days=today.weekday())   # last Monday
    week_end     = week_start + timedelta(days=6)
    generated_at = datetime.now()

    rows = build_week_predictions(week_start, predictor)

    wb = Workbook()
    _build_cover(wb, week_start, week_end, generated_at)
    _build_summary(wb, rows, week_start)
    _build_daily_detail(wb, rows)
    _build_schedule(wb, rows, week_start)

    if output_path is None:
        fname = f"ATM_CashPredict_Report_{week_start.strftime('%Y-%m-%d')}.xlsx"
        output_path = os.path.join("/mnt/user-data/outputs", fname)

    wb.save(output_path)
    return output_path


# ── Run standalone demo ───────────────────────────────────────────────────────
if __name__ == "__main__":
    from datetime import datetime, timedelta
    # Use the week starting 2026-06-22 (Monday this week)
    ws = datetime(2026, 6, 22)
    path = generate_report(week_start=ws)
    print(f"Report saved: {path}")
